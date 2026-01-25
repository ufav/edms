"""
Service for scheduled email sending of review Excel reports
"""
import json
import logging
from datetime import datetime, time, timedelta, timezone
from typing import List, Dict, Any, Optional
from sqlalchemy.orm import Session
from app.core.database import get_db
from app.models.user_settings import UserSettings
from app.models.user import User
from app.models.project import Project
from app.models.project_participant import ProjectParticipant
from app.services.email_service import email_service
import io
from urllib.parse import quote

logger = logging.getLogger(__name__)
# Устанавливаем уровень логирования для отладки
logger.setLevel(logging.DEBUG)
# Устанавливаем уровень логирования для отладки
logger.setLevel(logging.DEBUG)


def get_last_sent_time(schedule: Dict[str, Any], db: Session) -> Optional[datetime]:
    """
    Получить время последней отправки для расписания
    Используем поле lastSentAt в самом расписании или отдельную таблицу
    """
    last_sent_at_str = schedule.get('lastSentAt')
    if last_sent_at_str:
        try:
            return datetime.fromisoformat(last_sent_at_str.replace('Z', '+00:00'))
        except (ValueError, AttributeError):
            return None
    return None


def update_last_sent_time(schedule: Dict[str, Any], db: Session, user_id: int, sent_time: datetime):
    """
    Обновить время последней отправки для расписания
    """
    from app.models.user_settings import UserSettings
    
    # Находим настройку
    setting = db.query(UserSettings).filter(
        UserSettings.user_id == user_id,
        UserSettings.page == 'reviews_email_schedule',
        UserSettings.settings_key == 'schedules'
    ).first()
    
    if not setting:
        setting = db.query(UserSettings).filter(
            UserSettings.user_id == user_id,
            UserSettings.page == 'reviews',
            UserSettings.settings_key == 'reviewEmailScheduleSettings'
        ).first()
    
    if not setting or not setting.settings_value:
        return
    
    try:
        settings_data = json.loads(setting.settings_value)
        
        # Если это объект с ключом 'schedules', обновляем список
        if isinstance(settings_data, dict) and 'schedules' in settings_data:
            schedules = settings_data['schedules']
        elif isinstance(settings_data, list):
            schedules = settings_data
            settings_data = {'schedules': schedules}
        else:
            return
        
        # Находим нужное расписание и обновляем lastSentAt
        schedule_id = schedule.get('id')
        for sched in schedules:
            if sched.get('id') == schedule_id:
                sched['lastSentAt'] = sent_time.isoformat()
                break
        
        # Сохраняем обратно
        setting.settings_value = json.dumps(settings_data)
        db.commit()
        logger.info(f"Updated lastSentAt in database for schedule {schedule.get('id')}")
    except (json.JSONDecodeError, KeyError, AttributeError) as e:
        logger.error(f"Failed to update lastSentAt for schedule {schedule.get('id')}: {e}")


def get_user_schedules(db: Session, user_id: int) -> List[Dict[str, Any]]:
    """
    Get all email schedules for a user
    
    Args:
        db: Database session
        user_id: User ID
        
    Returns:
        List of schedule dictionaries
    """
    # Ищем настройки в двух возможных форматах:
    # 1. Старый формат: page='reviews', settings_key='reviewEmailScheduleSettings'
    # 2. Новый формат: page='reviews_email_schedule', settings_key='schedules'
    setting = db.query(UserSettings).filter(
        UserSettings.user_id == user_id,
        UserSettings.page == 'reviews_email_schedule',
        UserSettings.settings_key == 'schedules'
    ).first()
    
    # Если не найдено, пробуем старый формат
    if not setting:
        setting = db.query(UserSettings).filter(
            UserSettings.user_id == user_id,
            UserSettings.page == 'reviews',
            UserSettings.settings_key == 'reviewEmailScheduleSettings'
        ).first()
    
    if not setting:
        logger.debug(f"No settings found for user {user_id}")
        return []
    
    if not setting.settings_value:
        logger.debug(f"Empty settings_value for user {user_id}")
        return []
    
    try:
        settings_data = json.loads(setting.settings_value)
        
        # Если это объект с ключом 'schedules', извлекаем список
        if isinstance(settings_data, dict) and 'schedules' in settings_data:
            schedules = settings_data['schedules']
        elif isinstance(settings_data, list):
            schedules = settings_data
        else:
            logger.warning(f"Unexpected settings format for user {user_id}: {type(settings_data)}")
            print(f"Unexpected settings format for user {user_id}: {type(settings_data)}")
            return []
        
        logger.info(f"Loaded {len(schedules) if isinstance(schedules, list) else 0} schedule(s) for user {user_id}")
        print(f"Loaded {len(schedules) if isinstance(schedules, list) else 0} schedule(s) for user {user_id}")
        if isinstance(schedules, list):
            # Логируем содержимое расписаний для отладки
            for idx, sched in enumerate(schedules):
                logger.info(f"  Schedule {idx}: {json.dumps(sched, indent=2, ensure_ascii=False)}")
            return schedules
        return []
    except json.JSONDecodeError as e:
        logger.error(f"Failed to parse schedules for user {user_id}: {e}")
        print(f"Failed to parse schedules for user {user_id}: {e}")
        return []


def should_send_now(schedule: Dict[str, Any], current_time: datetime, db: Session = None) -> bool:
    """
    Check if schedule should trigger now based on current time
    
    Args:
        schedule: Schedule dictionary with 'schedule', 'time', 'timezone', 'dayOfWeek', 'dayOfMonth'
        current_time: Current datetime in UTC
        db: Optional database session for checking last sent time
        
    Returns:
        True if schedule should trigger now
    """
    # Проверяем, не отправляли ли мы уже в эту же минуту (защита от повторных отправок)
    if db:
        last_sent = get_last_sent_time(schedule, db)
        if last_sent:
            # Вычисляем разницу во времени
            time_diff = current_time - last_sent
            
            # Если отправляли менее 2 минут назад - точно не отправляем (защита от дублей)
            if time_diff.total_seconds() < 120:  # 2 минуты
                schedule_id = schedule.get('id', 'unknown')
                logger.info(
                    f"Schedule {schedule_id}: Already sent {time_diff.total_seconds():.0f} seconds ago, skipping to prevent duplicates"
                )
                print(f"Schedule {schedule_id}: Already sent {time_diff.total_seconds():.0f} seconds ago, skipping")
                return False
            
            schedule_type = schedule.get('schedule')
            # Для daily - не отправляем повторно в течение часа
            if schedule_type == 'daily':
                if time_diff.total_seconds() < 3600:  # 1 час
                    logger.debug(f"Schedule {schedule.get('id')}: Already sent within last hour, skipping")
                    return False
            # Для weekly - не отправляем повторно в течение дня
            elif schedule_type == 'weekly':
                if time_diff.total_seconds() < 86400:  # 1 день
                    logger.debug(f"Schedule {schedule.get('id')}: Already sent today, skipping")
                    return False
            # Для monthly - не отправляем повторно в течение дня
            elif schedule_type == 'monthly':
                if time_diff.total_seconds() < 86400:  # 1 день
                    logger.debug(f"Schedule {schedule.get('id')}: Already sent today, skipping")
                    return False
    schedule_type = schedule.get('schedule')
    schedule_time_str = schedule.get('time', '09:00')
    schedule_timezone = schedule.get('timezone', 'UTC+0')
    
    # Parse timezone offset (e.g., "UTC+5" -> 5, "UTC-3" -> -3)
    try:
        tz_offset_str = schedule_timezone.replace('UTC', '').strip()
        if tz_offset_str.startswith('+'):
            tz_offset = int(tz_offset_str[1:])
        elif tz_offset_str.startswith('-'):
            tz_offset = -int(tz_offset_str[1:])
        else:
            # Если нет знака, пробуем распарсить как число
            tz_offset = int(tz_offset_str) if tz_offset_str else 0
    except (ValueError, AttributeError):
        logger.warning(f"Invalid timezone format: {schedule_timezone}, using UTC+0")
        tz_offset = 0
    
    # Преобразуем текущее UTC время в локальное время пользователя
    # Если пользователь в UTC+3, то его локальное время = UTC + 3 часа
    schedule_local_time = current_time + timedelta(hours=tz_offset)
    
    # Парсим время из расписания (локальное время пользователя)
    # Поддерживаем форматы "18:45" и "18-45"
    try:
        # Пробуем сначала с двоеточием, потом с дефисом
        if ':' in schedule_time_str:
            hour, minute = map(int, schedule_time_str.split(':'))
        elif '-' in schedule_time_str:
            hour, minute = map(int, schedule_time_str.split('-'))
        else:
            raise ValueError(f"Unknown time format: {schedule_time_str}")
        schedule_time = time(hour, minute)
    except (ValueError, AttributeError) as e:
        logger.warning(f"Invalid time format: {schedule_time_str}, using 09:00. Error: {e}")
        schedule_time = time(9, 0)
    
    # Получаем текущее локальное время пользователя
    current_local_time = schedule_local_time.time()
    current_local_date = schedule_local_time.date()
    
    # Проверяем, совпадает ли время (проверяем только нужную минуту, защита от дублей через lastSentAt)
    # Вычисляем разницу в минутах
    current_minutes = current_local_time.hour * 60 + current_local_time.minute
    schedule_minutes = schedule_time.hour * 60 + schedule_time.minute
    time_diff_minutes = abs(current_minutes - schedule_minutes)
    
    # Должна быть нужная минута (разница 0)
    # Не ограничиваем секунды - если мы в нужную минуту, отправляем
    # Защита от повторных отправок обеспечивается проверкой lastSentAt выше
    if time_diff_minutes != 0:
        schedule_id = schedule.get('id', 'unknown')
        logger.debug(
            f"Schedule {schedule_id}: Time mismatch - wrong minute. "
            f"Current local time: {current_local_time}, Schedule time: {schedule_time}, "
            f"UTC time: {current_time}, Timezone: {schedule_timezone}, Diff: {time_diff_minutes} min"
        )
        return False
    
    # Проверяем тип расписания
    if schedule_type == 'daily':
        logger.info(
            f"Schedule {schedule.get('id')}: Daily schedule triggered. "
            f"Local time: {current_local_time}, UTC time: {current_time}, Timezone: {schedule_timezone}"
        )
        return True
    
    elif schedule_type == 'weekly':
        day_of_week = schedule.get('dayOfWeek', 1)  # 1 = Monday
        # Python weekday: Monday=0, Sunday=6
        # Наш dayOfWeek: Monday=1, Sunday=7
        current_weekday = schedule_local_time.weekday() + 1
        matches = current_weekday == day_of_week
        if matches:
            logger.info(
                f"Schedule {schedule.get('id')}: Weekly schedule triggered. "
                f"Day: {current_weekday}, Local time: {current_local_time}, UTC time: {current_time}, Timezone: {schedule_timezone}"
            )
        return matches
    
    elif schedule_type == 'monthly':
        day_of_month = schedule.get('dayOfMonth', 1)
        current_day = schedule_local_time.day
        matches = current_day == day_of_month
        if matches:
            logger.info(
                f"Schedule {schedule.get('id')}: Monthly schedule triggered. "
                f"Day: {current_day}, Local time: {current_local_time}, UTC time: {current_time}, Timezone: {schedule_timezone}"
            )
        return matches
    
    return False


def get_recipient_email(schedule: Dict[str, Any], db: Session, project_id: Optional[int] = None) -> Optional[str]:
    """
    Get email address for schedule recipient
    
    Args:
        schedule: Schedule dictionary
        db: Database session
        project_id: Optional project ID for company/user recipients
        
    Returns:
        Email address or None
    """
    recipient_type = schedule.get('recipientType')
    
    if recipient_type == 'self':
        return schedule.get('userEmail')
    
    elif recipient_type == 'company':
        contact_email = schedule.get('contactEmail')
        if contact_email:
            return contact_email
        
        # Fallback: get from project participant
        if project_id:
            company_id = schedule.get('companyId')
            if company_id:
                participant = db.query(ProjectParticipant).filter(
                    ProjectParticipant.project_id == project_id,
                    ProjectParticipant.company_id == company_id
                ).first()
                if participant and participant.contact_email:
                    return participant.contact_email
    
    elif recipient_type == 'user':
        # Extract email from userName if present (format: "Name (email@example.com)")
        user_name = schedule.get('userName', '')
        if '(' in user_name and ')' in user_name:
            email_start = user_name.rfind('(') + 1
            email_end = user_name.rfind(')')
            return user_name[email_start:email_end]
        
        # Or get from user ID
        user_id = schedule.get('userId')
        if user_id:
            user = db.query(User).filter(User.id == user_id).first()
            if user and user.email:
                return user.email
    
    return None


def process_single_schedule(schedule: Dict[str, Any], current_time: datetime):
    """
    Обрабатывает одно расписание отправки письма
    
    Args:
        schedule: Словарь с настройками расписания
        current_time: Текущее время в UTC
    """
    from app.core.database import SessionLocal
    
    # Создаем отдельную сессию БД для каждого расписания
    db = SessionLocal()
    try:
        # Get project ID from schedule
        project_id = schedule.get('projectId')
        if not project_id:
            logger.warning(f"No project_id in schedule {schedule.get('id')}")
            return
        
        # Check only_overdue filter
        only_overdue = schedule.get('onlyOverdue', False)
        
        # Определяем фильтр по компании, если получатель - компания
        selected_company = None
        is_company_recipient = schedule.get('recipientType') == 'company'
        if is_company_recipient:
            # Получаем название компании из расписания
            company_name = schedule.get('companyName')
            if company_name:
                selected_company = company_name
        
        # Get recipient email
        recipient_email = get_recipient_email(schedule, db, project_id)
        if not recipient_email:
            logger.warning(f"No email found for schedule {schedule.get('id')}")
            return
        
        # Generate Excel file
        excel_bytes = generate_review_excel_bytes(
            db=db,
            project_id=project_id,
            search=None,
            selected_company=selected_company,
            only_overdue=only_overdue,
            language=schedule.get('language', 'ru'),
            exclude_awaiting_company=is_company_recipient
        )
        
        if not excel_bytes:
            logger.warning(f"Failed to generate Excel for schedule {schedule.get('id')}")
            return
        
        # Get project name
        project = db.query(Project).filter(Project.id == project_id).first()
        project_name = project.name if project else "Unknown Project"
        
        # Generate filename
        filename = f"reviews_{project_name}_{current_time.strftime('%Y-%m-%d')}.xlsx"
        
        # Send email
        success = email_service.send_review_excel_report(
            to_emails=[recipient_email],
            project_name=project_name,
            excel_content=excel_bytes,
            filename=filename,
            language=schedule.get('language', 'ru')
        )
        
        if success:
            logger.info(f"Sent review report email to {recipient_email} for schedule {schedule.get('id')}")
            # Обновляем время последней отправки СРАЗУ после успешной отправки
            # Это важно для предотвращения повторных отправок
            # Получаем user_id из расписания
            # Для 'self' - берем из расписания, для других - находим по email
            user_id_to_update = None
            if schedule.get('recipientType') == 'self':
                # Для 'self' нужно найти user_id по email из расписания
                user_email = schedule.get('userEmail')
                if user_email:
                    user = db.query(User).filter(User.email == user_email).first()
                    if user:
                        user_id_to_update = user.id
            else:
                # Для company/user - находим по recipient_email
                user = db.query(User).filter(User.email == recipient_email).first()
                if user:
                    user_id_to_update = user.id
            
            if user_id_to_update:
                try:
                    update_last_sent_time(schedule, db, user_id_to_update, current_time)
                    logger.info(f"Updated lastSentAt for schedule {schedule.get('id')} to {current_time.isoformat()}")
                except Exception as e:
                    logger.error(f"Failed to update lastSentAt for schedule {schedule.get('id')}: {e}")
        else:
            logger.error(f"Failed to send review report email to {recipient_email} for schedule {schedule.get('id')}")
            
    except Exception as e:
        logger.error(f"Error processing schedule {schedule.get('id')}: {e}", exc_info=True)
    finally:
        db.close()


def process_scheduled_emails(db: Session):
    """
    Process all scheduled email sends
    
    This function should be called periodically (e.g., every minute via cron or scheduler)
    Обрабатывает расписания параллельно для разных пользователей
    """
    from concurrent.futures import ThreadPoolExecutor, as_completed
    
    current_time = datetime.now(timezone.utc)
    logger.info(f"=== process_scheduled_emails called at {current_time} ===")
    print(f"=== process_scheduled_emails called at {current_time} ===")
    
    # Get all users with review email schedules
    # Ищем в обоих форматах: старом и новом
    users_with_schedules = db.query(User).join(
        UserSettings,
        UserSettings.user_id == User.id
    ).filter(
        (
            (UserSettings.page == 'reviews_email_schedule') & (UserSettings.settings_key == 'schedules')
        ) | (
            (UserSettings.page == 'reviews') & (UserSettings.settings_key == 'reviewEmailScheduleSettings')
        )
    ).all()
    
    logger.info(f"Found {len(users_with_schedules)} user(s) with email schedules")
    print(f"Found {len(users_with_schedules)} user(s) with email schedules")
    
    # Собираем все расписания, которые нужно обработать
    schedules_to_process = []
    schedules_by_user = {}  # Для логирования
    
    for user in users_with_schedules:
        schedules = get_user_schedules(db, user.id)
        user_schedules_count = 0
        
        for schedule in schedules:
            # Логируем информацию о проверяемом расписании для отладки
            schedule_id = schedule.get('id', 'unknown')
            schedule_time = schedule.get('time', 'N/A')
            schedule_type = schedule.get('schedule', 'N/A')
            schedule_tz = schedule.get('timezone', 'N/A')
            
            logger.debug(
                f"Checking schedule {schedule_id} for user {user.id}: "
                f"type={schedule_type}, time={schedule_time}, tz={schedule_tz}, "
                f"current_utc={current_time}"
            )
            
            # Проверяем расписание с передачей db для проверки lastSentAt
            if should_send_now(schedule, current_time, db):
                # Дополнительная проверка: если уже есть в списке на обработку - пропускаем
                # (защита от дублей в одном цикле)
                schedule_already_queued = any(
                    s.get('id') == schedule_id for s in schedules_to_process
                )
                if schedule_already_queued:
                    logger.warning(f"Schedule {schedule_id} already queued for processing, skipping duplicate")
                    continue
                
                schedules_to_process.append(schedule)
                user_schedules_count += 1
                logger.info(
                    f"Schedule {schedule_id} matched! Will process. "
                    f"Time: {schedule_time}, Type: {schedule_type}, TZ: {schedule_tz}"
                )
        
        if user_schedules_count > 0:
            schedules_by_user[user.id] = user_schedules_count
    
    # Логируем общую информацию о проверке
    if users_with_schedules:
        logger.info(
            f"Checked {len(users_with_schedules)} user(s) with schedules. "
            f"Found {len(schedules_to_process)} schedule(s) to process. "
            f"Current UTC time: {current_time}"
        )
    
    if not schedules_to_process:
        logger.info(f"No schedules to process at {current_time}")
        print(f"No schedules to process at {current_time}")
        return
    
    # Логируем информацию о расписаниях
    logger.info(
        f"Processing {len(schedules_to_process)} scheduled email(s) in parallel. "
        f"Users with schedules: {len(schedules_by_user)}. "
        f"Multiple schedules per user: {sum(1 for count in schedules_by_user.values() if count > 1)}"
    )
    
    # Обрабатываем расписания параллельно
    # Используем ThreadPoolExecutor для параллельной обработки
    # Максимум 10 одновременных задач, чтобы не перегрузить систему
    max_workers = min(10, len(schedules_to_process))
    
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        # Запускаем обработку всех расписаний
        future_to_schedule = {
            executor.submit(process_single_schedule, schedule, current_time): schedule
            for schedule in schedules_to_process
        }
        
        # Ждем завершения всех задач
        for future in as_completed(future_to_schedule):
            schedule = future_to_schedule[future]
            try:
                future.result()  # Получаем результат (или исключение, если было)
            except Exception as e:
                logger.error(f"Unexpected error in schedule {schedule.get('id')}: {e}", exc_info=True)
    
    logger.info(f"Completed processing {len(schedules_to_process)} scheduled email(s)")


def generate_review_excel_bytes(
    db: Session,
    project_id: Optional[int] = None,
    search: Optional[str] = None,
    selected_company: Optional[str] = None,
    only_overdue: Optional[bool] = False,
    language: str = "ru",
    exclude_awaiting_company: bool = False
) -> Optional[bytes]:
    """
    Generate Excel file bytes for reviews export
    This is a refactored version of export_reviews_to_excel endpoint logic
    
    Args:
        db: Database session
        project_id: Optional project ID
        search: Optional search string
        selected_company: Optional company filter (company name)
        only_overdue: Filter only overdue reviews
        language: Language for export
        
    Returns:
        Excel file bytes or None
    """
    try:
        # Import here to avoid circular imports
        from app.models.references import WorkflowStatus
        from app.models.document import Document, DocumentRevision
        from app.models.document import File as FileModel
        from app.models.project import Project, WorkflowPresetSequence
        from app.models.references import RevisionStep, RevisionDescription
        from app.models.transmittal import Transmittal, TransmittalRevision
        from app.models.references import Company, TransmittalStatus
        from sqlalchemy import func, and_
        from collections import defaultdict
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        # Get "In Review" status
        in_review_status = db.query(WorkflowStatus).filter(
            WorkflowStatus.name == "In Review"
        ).first()
        
        if not in_review_status:
            return None
        
        # Create subquery for latest revision
        latest_revision_subquery = db.query(
            DocumentRevision.document_id,
            func.max(DocumentRevision.created_at).label('max_created_at')
        ).group_by(DocumentRevision.document_id).subquery()
        
        # Build query
        query = db.query(
            Document,
            DocumentRevision,
            Project,
            WorkflowPresetSequence
        ).outerjoin(
            latest_revision_subquery,
            Document.id == latest_revision_subquery.c.document_id
        ).outerjoin(
            DocumentRevision,
            and_(
                DocumentRevision.document_id == Document.id,
                DocumentRevision.created_at == latest_revision_subquery.c.max_created_at
            )
        ).join(
            Project,
            Project.id == Document.project_id
        ).outerjoin(
            WorkflowPresetSequence,
            and_(
                WorkflowPresetSequence.preset_id == Project.workflow_preset_id,
                WorkflowPresetSequence.revision_step_id == DocumentRevision.revision_step_id,
                WorkflowPresetSequence.revision_description_id == DocumentRevision.revision_description_id
            )
        ).filter(
            Document.is_deleted == 0,
            DocumentRevision.workflow_status_id == in_review_status.id
        )
        
        if project_id:
            query = query.filter(Project.id == project_id)
        
        results = query.order_by(Document.updated_at.desc()).all()
        
        # Load related data (similar to export_reviews_to_excel)
        revision_step_ids = []
        revision_description_ids = []
        
        for row in results:
            doc, revision, project, sequence = row
            if revision and revision.revision_step_id:
                revision_step_ids.append(revision.revision_step_id)
            if revision and revision.revision_description_id:
                revision_description_ids.append(revision.revision_description_id)
        
        revision_steps = {}
        if revision_step_ids:
            for rs in db.query(RevisionStep).filter(RevisionStep.id.in_(revision_step_ids)).all():
                revision_steps[rs.id] = rs
        
        revision_descriptions = {}
        if revision_description_ids:
            for rd in db.query(RevisionDescription).filter(RevisionDescription.id.in_(revision_description_ids)).all():
                revision_descriptions[rd.id] = rd
        
        revision_ids = [row[1].id for row in results if row[1] is not None]
        files_by_revision_id = defaultdict(list)
        if revision_ids:
            all_files = db.query(FileModel).filter(
                FileModel.revision_id.in_(revision_ids),
                FileModel.is_deleted == 0
            ).all()
            for f in all_files:
                files_by_revision_id[f.revision_id].append(f)
        
        # Load company information
        awaiting_company_by_revision_id = {}
        if revision_ids:
            sent_status = db.query(TransmittalStatus).filter(
                TransmittalStatus.name == "Sent"
            ).first()
            
            if sent_status:
                transmittal_data = db.query(
                    TransmittalRevision.revision_id,
                    Company.id,
                    Company.name,
                    Company.name_native
                ).join(
                    Transmittal, Transmittal.id == TransmittalRevision.transmittal_id
                ).join(
                    Company, Company.id == Transmittal.counterparty_id
                ).filter(
                    TransmittalRevision.revision_id.in_(revision_ids),
                    Transmittal.direction == 'out',
                    Transmittal.status_id == sent_status.id,
                    Transmittal.is_deleted == 0
                ).all()
                
                for rev_id, company_id, company_name, company_name_native in transmittal_data:
                    awaiting_company_by_revision_id[rev_id] = {
                        "id": company_id,
                        "name": company_name,
                        "name_native": company_name_native
                    }
        
        # Формируем данные для Excel
        excel_rows = []
        now = datetime.now(timezone.utc)
        
        # Вспомогательные функции для форматирования
        def format_date(date_obj):
            if not date_obj:
                return ''
            if isinstance(date_obj, str):
                try:
                    date_obj = datetime.fromisoformat(date_obj.replace('Z', '+00:00'))
                except:
                    return ''
            return date_obj.strftime('%d.%m.%Y %H:%M')
        
        def format_file_size(bytes_size):
            if not bytes_size:
                return '0 B'
            k = 1024
            sizes = ['B', 'KB', 'MB', 'GB']
            i = 0
            size = float(bytes_size)
            while size >= k and i < len(sizes) - 1:
                size /= k
                i += 1
            return f"{round(size * 100) / 100} {sizes[i]}"
        
        # Локализация заголовков
        headers_ru = {
            "document": "Документ",
            "title": "Название",
            "project": "Проект",
            "revision": "Ревизия",
            "current_step": "Текущий шаг",
            "awaiting_company": "Ожидается от",
            "release_date": "Дата выпуска",
            "due_days": "Срок (дней)",
            "due_date": "Срок выполнения",
            "overdue": "Просрочено",
            "file": "Файл",
            "size": "Размер"
        }
        
        headers_en = {
            "document": "Document",
            "title": "Title",
            "project": "Project",
            "revision": "Revision",
            "current_step": "Current Step",
            "awaiting_company": "Awaiting Company",
            "release_date": "Release Date",
            "due_days": "Due Days",
            "due_date": "Due Date",
            "overdue": "Overdue",
            "file": "File",
            "size": "Size"
        }
        
        headers = headers_ru if language == "ru" else headers_en
        
        for row in results:
            doc, revision, project, sequence = row
            
            # Применяем фильтры
            if search:
                search_lower = search.lower()
                if not (search_lower in (doc.title or "").lower() or 
                       search_lower in (doc.number or "").lower() or
                       search_lower in (project.name or "").lower()):
                    continue
            
            if only_overdue:
                if revision and revision.created_at and sequence and sequence.due_days:
                    due_date = revision.created_at + timedelta(days=sequence.due_days)
                    if due_date >= now:
                        continue
                else:
                    continue
            
            if selected_company:
                if selected_company == '__internal__':
                    if sequence and sequence.requires_transmittal != False:
                        continue
                else:
                    awaiting_company = awaiting_company_by_revision_id.get(revision.id if revision else None)
                    if not awaiting_company or awaiting_company.get("name") != selected_company:
                        continue
            
            # Получаем информацию о шаге и описании
            step_info = None
            description_info = None
            
            if revision:
                step = revision_steps.get(revision.revision_step_id)
                description = revision_descriptions.get(revision.revision_description_id)
                
                step_info = {
                    "id": step.id if step else None,
                    "code": step.code if step else None,
                    "description": step.description if step else None,
                    "description_native": step.description_native if step else None
                } if step else None
                
                description_info = {
                    "id": description.id if description else None,
                    "code": description.code if description else None,
                    "description": description.description if description else None,
                    "description_native": description.description_native if description else None
                } if description else None
            
            files_info = files_by_revision_id.get(revision.id, []) if revision else []
            
            # Вычисляем due_date и is_overdue
            due_date = None
            is_overdue = False
            
            if revision and revision.created_at and sequence and sequence.due_days:
                due_date = revision.created_at + timedelta(days=sequence.due_days)
                is_overdue = due_date < now
            
            # Формируем поле просрочено
            overdue_value = ''
            if is_overdue and due_date:
                overdue_days = (now - due_date).days
                if overdue_days > 0:
                    if language == "ru":
                        # Склонение для русского
                        last_digit = overdue_days % 10
                        last_two = overdue_days % 100
                        if 11 <= last_two <= 14:
                            days_word = "дней"
                        elif last_digit == 1:
                            days_word = "день"
                        elif 2 <= last_digit <= 4:
                            days_word = "дня"
                        else:
                            days_word = "дней"
                        overdue_value = f"Просрочено на {overdue_days} {days_word}"
                    else:
                        days_word = "day" if overdue_days == 1 else "days"
                        overdue_value = f"Overdue by {overdue_days} {days_word}"
                else:
                    overdue_value = "Просрочено" if language == "ru" else "Overdue"
            
            # Формируем поле текущего шага
            current_step_value = ''
            if step_info:
                step_desc = (step_info.get("description_native") if language == "ru" else step_info.get("description")) or step_info.get("description") or ''
                if step_desc and step_desc.strip():
                    current_step_value = f"{step_info.get('code')} - {step_desc}"
                else:
                    current_step_value = step_info.get('code') or ''
            
            # Формируем поле ревизии
            revision_value = ''
            if description_info:
                revision_value = f"{description_info.get('code')}{revision.number if revision else ''}"
            elif revision:
                revision_value = revision.number or ''
            
            # Формируем строку данных
            row_data = [
                doc.number or '',
                doc.title or '',
                project.name if project else '',
                revision_value,
                current_step_value,
            ]
            
            # Добавляем поле "Ожидается от" только если не исключаем его
            if not exclude_awaiting_company:
                awaiting_company_value = ''
                awaiting_company = awaiting_company_by_revision_id.get(revision.id if revision else None)
                if awaiting_company:
                    awaiting_company_value = awaiting_company.get("name") or ''
                elif sequence and sequence.requires_transmittal == False:
                    awaiting_company_value = "Внутреннее ревью" if language == "ru" else "Internal Review"
                row_data.append(awaiting_company_value)
            
            # Остальные поля
            row_data.extend([
                format_date(revision.created_at if revision else None),
                sequence.due_days if sequence else '',
                format_date(due_date) if due_date else '',
                overdue_value,
                files_info[0].file_name if files_info and len(files_info) > 0 else '',
                format_file_size(files_info[0].file_size if files_info and len(files_info) > 0 else None)
            ])
            
            excel_rows.append(row_data)
        
        # Создаем Excel файл
        wb = Workbook()
        ws = wb.active
        ws.title = "Reviews" if language == "en" else "Ревью"
        
        # Добавляем заголовки
        header_row = [
            headers["document"],
            headers["title"],
            headers["project"],
            headers["revision"],
            headers["current_step"],
        ]
        
        # Добавляем заголовок "Ожидается от" только если не исключаем его
        if not exclude_awaiting_company:
            header_row.append(headers["awaiting_company"])
        
        header_row.extend([
            headers["release_date"],
            headers["due_days"],
            headers["due_date"],
            headers["overdue"],
            headers["file"],
            headers["size"]
        ])
        
        ws.append(header_row)
        
        # Применяем стили к заголовкам
        header_fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Добавляем данные
        data_alignment = Alignment(vertical="center", wrap_text=True)
        data_border = Border(
            left=Side(style='thin', color='D0D0D0'),
            right=Side(style='thin', color='D0D0D0'),
            top=Side(style='thin', color='D0D0D0'),
            bottom=Side(style='thin', color='D0D0D0')
        )
        
        # Создаем стиль для просроченных записей (красный текст)
        overdue_font = Font(color="FF0000", bold=False)
        
        # Определяем индекс колонки "Overdue" из заголовков
        overdue_col_index = None
        for idx, header in enumerate(header_row, start=1):
            if header == headers["overdue"]:
                overdue_col_index = idx
                break
        
        for row_data in excel_rows:
            ws.append(row_data)
            for col_idx, cell in enumerate(ws[ws.max_row], start=1):
                cell.alignment = data_alignment
                cell.border = data_border
                
                # Применяем красный цвет к колонке "Overdue", если значение не пустое
                if overdue_col_index and col_idx == overdue_col_index:
                    cell_value = str(cell.value) if cell.value else ''
                    if cell_value and (cell_value != ''):
                        cell.font = overdue_font
        
        # Настраиваем ширину колонок
        # Ширины: document, title, project, revision, current_step, [awaiting_company], release_date, due_days, due_date, overdue, file, size
        column_widths = [15, 30, 20, 12, 25]
        if not exclude_awaiting_company:
            column_widths.append(20)  # awaiting_company
        column_widths.extend([20, 12, 20, 20, 25, 12])  # release_date, due_days, due_date, overdue, file, size
        
        for idx, width in enumerate(column_widths, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = width
        
        # Устанавливаем высоту строки для заголовков
        ws.row_dimensions[1].height = 25
        
        # Замораживаем первую строку
        ws.freeze_panes = 'A2'
        
        # Добавляем автофильтр для заголовков
        if excel_rows:
            last_col = get_column_letter(len(header_row))
            ws.auto_filter.ref = f"A1:{last_col}{len(excel_rows) + 1}"
        
        # Сохраняем в BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.read()
        
    except Exception as e:
        logger.error(f"Error generating Excel bytes: {e}", exc_info=True)
        return None
