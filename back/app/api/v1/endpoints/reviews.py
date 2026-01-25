"""
API endpoints for document reviews and approvals
"""

from fastapi import APIRouter, Depends, HTTPException, Request, Query
from fastapi.responses import Response
from sqlalchemy.orm import Session
from sqlalchemy import func, and_, or_
from typing import List, Optional
from pydantic import BaseModel
from datetime import datetime, timedelta, timezone
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from urllib.parse import quote

class ApproveRequest(BaseModel):
    comments: Optional[str] = None

class RejectRequest(BaseModel):
    comments: Optional[str] = None

from app.core.database import get_db
from app.models.user import User
from app.models.document import Document, DocumentRevision
from app.models.document import File as FileModel
from app.models.project import Project, WorkflowPresetSequence, ProjectStatusEnum, ProjectStatusEnum
from app.models.references import WorkflowStatus, RevisionStep, RevisionDescription
from app.api.v1.endpoints.auth import get_current_user
from app.services.audit_service import log_action

router = APIRouter()

def _check_and_update_project_status_on_approval(project_id: int, db: Session):
    """
    Проверяет, есть ли утвержденные документы в проекте.
    Если это первый утвержденный документ, обновляет статус проекта на ACTIVE.
    """
    if not project_id:
        return
    
    # Получаем проект
    project = db.query(Project).filter(Project.id == project_id, Project.is_deleted == 0).first()
    if not project:
        return
    
    # Если статус проекта уже ACTIVE или выше, ничего не делаем
    if project.status == ProjectStatusEnum.ACTIVE or project.status == ProjectStatusEnum.COMPLETED:
        return
    
    # Получаем ID утвержденных статусов
    approved_like_statuses = db.query(WorkflowStatus).filter(
        WorkflowStatus.name.in_(["Approved", "Approved with Comments", "Not Reviewed"])
    ).all()
    approved_like_ids = {s.id for s in approved_like_statuses}
    
    if not approved_like_ids:
        return
    
    # Проверяем, есть ли в проекте документы с утвержденными ревизиями
    # Ищем документы проекта, у которых последняя ревизия в утвержденном статусе
    from sqlalchemy import and_
    
    # Получаем все документы проекта
    project_documents = db.query(Document).filter(
        Document.project_id == project_id,
        Document.is_deleted == 0
    ).all()
    
    if not project_documents:
        return
    
    # Получаем последние ревизии для всех документов одним запросом (устранение N+1)
    document_ids = [doc.id for doc in project_documents]
    
    # Создаем подзапрос для получения последней ревизии каждого документа
    from sqlalchemy import func
    latest_revision_subquery = db.query(
        DocumentRevision.document_id,
        func.max(DocumentRevision.created_at).label('max_created_at')
    ).filter(
        DocumentRevision.document_id.in_(document_ids),
        DocumentRevision.is_deleted == 0
    ).group_by(DocumentRevision.document_id).subquery()
    
    # Получаем все последние ревизии одним запросом
    latest_revisions = db.query(DocumentRevision).join(
        latest_revision_subquery,
        and_(
            DocumentRevision.document_id == latest_revision_subquery.c.document_id,
            DocumentRevision.created_at == latest_revision_subquery.c.max_created_at,
            DocumentRevision.is_deleted == 0
        )
    ).all()
    
    # Проверяем, есть ли хотя бы одна ревизия в утвержденном статусе
    has_approved_document = any(
        rev.workflow_status_id in approved_like_ids 
        for rev in latest_revisions
    )
    
    # Если есть утвержденный документ и статус проекта PLANNING, обновляем на ACTIVE
    if has_approved_document and project.status == ProjectStatusEnum.PLANNING:
        project.status = ProjectStatusEnum.ACTIVE
        db.commit()
        print(f"Project {project_id} status updated to ACTIVE - first approved document found")

@router.get("/")
async def get_reviews(
    project_id: int = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    # Временно возвращаем пустой список
    return []


@router.get("/pending-approvals")
async def get_pending_approvals(
    skip: int = 0,
    limit: int = 100,
    project_id: int = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Получить документы, ожидающие внутреннего утверждения (без трансмиттала)
    """
    # Получаем статус "In Review"
    in_review_status = db.query(WorkflowStatus).filter(
        WorkflowStatus.name == "In Review"
    ).first()
    
    if not in_review_status:
        return []
    
    # Создаем подзапрос для получения последней ревизии каждого документа
    latest_revision_subquery = db.query(
        DocumentRevision.document_id,
        func.max(DocumentRevision.created_at).label('max_created_at')
    ).group_by(DocumentRevision.document_id).subquery()
    
    # Запрос с информацией о последовательности для определения доступности кнопок
    query = db.query(
        Document,
        DocumentRevision,
        Project,
        WorkflowPresetSequence
    ).outerjoin(
        latest_revision_subquery,
        Document.id == latest_revision_subquery.c.document_id
    ).outerjoin(
        DocumentRevision,
        and_(
            DocumentRevision.document_id == Document.id,
            DocumentRevision.created_at == latest_revision_subquery.c.max_created_at
        )
    ).join(
        Project,
        Project.id == Document.project_id
    ).outerjoin(
        WorkflowPresetSequence,
        and_(
            WorkflowPresetSequence.preset_id == Project.workflow_preset_id,
            WorkflowPresetSequence.revision_step_id == DocumentRevision.revision_step_id,
            WorkflowPresetSequence.revision_description_id == DocumentRevision.revision_description_id
        )
    ).filter(
        Document.is_deleted == 0,
        DocumentRevision.workflow_status_id == in_review_status.id,
        Project.id == project_id
    )
    
    # Выполняем запрос с пагинацией
    results = query.order_by(Document.updated_at.desc()).offset(skip).limit(limit).all()
    
    # Получаем все нужные ID для batch запросов
    revision_step_ids = []
    revision_description_ids = []
    
    for row in results:
        doc, revision, project, sequence = row
        if revision and revision.revision_step_id:
            revision_step_ids.append(revision.revision_step_id)
        if revision and revision.revision_description_id:
            revision_description_ids.append(revision.revision_description_id)
    
    # Загружаем все RevisionStep одним запросом
    revision_steps = {}
    if revision_step_ids:
        for rs in db.query(RevisionStep).filter(RevisionStep.id.in_(revision_step_ids)).all():
            revision_steps[rs.id] = rs
    
    # Загружаем все RevisionDescription одним запросом
    revision_descriptions = {}
    if revision_description_ids:
        for rd in db.query(RevisionDescription).filter(RevisionDescription.id.in_(revision_description_ids)).all():
            revision_descriptions[rd.id] = rd
    
    # Пакетно загружаем файлы для всех ревизий (устранение N+1)
    from collections import defaultdict
    revision_ids = [row[1].id for row in results if row[1] is not None]
    files_by_revision_id = defaultdict(list)
    if revision_ids:
        all_files = db.query(FileModel).filter(
            FileModel.revision_id.in_(revision_ids),
            FileModel.is_deleted == 0
        ).all()
        for f in all_files:
            files_by_revision_id[f.revision_id].append(f)
    
    # Загружаем информацию о компаниях-получателях из исходящих трансмитталов (ожидаем ответ)
    from app.models.transmittal import Transmittal, TransmittalRevision
    from app.models.references import Company, TransmittalStatus
    
    awaiting_company_by_revision_id = {}
    if revision_ids:
        # Получаем статус "Sent" (отправлен, но ещё не получен ответ)
        sent_status = db.query(TransmittalStatus).filter(
            TransmittalStatus.name == "Sent"
        ).first()
        
        if sent_status:
            # Ищем исходящие трансмитталы (direction='out') в статусе "Sent" для ревизий
            transmittal_data = db.query(
                TransmittalRevision.revision_id,
                Company.id,
                Company.name,
                Company.name_native
            ).join(
                Transmittal, Transmittal.id == TransmittalRevision.transmittal_id
            ).join(
                Company, Company.id == Transmittal.counterparty_id
            ).filter(
                TransmittalRevision.revision_id.in_(revision_ids),
                Transmittal.direction == 'out',
                Transmittal.status_id == sent_status.id,
                Transmittal.is_deleted == 0
            ).all()
            
            for rev_id, company_id, company_name, company_name_native in transmittal_data:
                awaiting_company_by_revision_id[rev_id] = {
                    "id": company_id,
                    "name": company_name,
                    "name_native": company_name_native
                }

    # Формируем результат
    result = []
    for row in results:
        doc, revision, project, sequence = row
        
        # Получаем информацию о шаге и описании
        step_info = None
        description_info = None
        
        if revision:
            step = revision_steps.get(revision.revision_step_id)
            description = revision_descriptions.get(revision.revision_description_id)
            
            step_info = {
                "id": step.id if step else None,
                "code": step.code if step else None,
                "description": step.description if step else None,
                "description_native": step.description_native if step else None
            } if step else None
            
            description_info = {
                "id": description.id if description else None,
                "code": description.code if description else None,
                "description": description.description if description else None,
                "description_native": description.description_native if description else None
            } if description else None
        
        # Берем файлы из заранее загруженной мапы
        files_info = files_by_revision_id.get(revision.id, []) if revision else []
        
        # Вычисляем due_date и is_overdue
        due_date = None
        is_overdue = False
        now = datetime.now(timezone.utc)
        
        if revision and revision.created_at and sequence and sequence.due_days:
            due_date = revision.created_at + timedelta(days=sequence.due_days)
            is_overdue = due_date < now
        
        result.append({
            "document_id": doc.id,
            "document_title": doc.title,
            "document_number": doc.number,
            "project_id": doc.project_id,
            "project_name": project.name if project else None,
            "revision_id": revision.id if revision else None,
            "revision_number": revision.number if revision else None,
            "files": [
                {
                    "id": f.id,
                    "file_name": f.file_name,
                    "file_size": f.file_size,
                    "file_type": f.file_type,
                }
                for f in files_info
            ],
            # Обратная совместимость - берем первый файл для старых полей
            "file_name": files_info[0].file_name if files_info and len(files_info) > 0 else None,
            "file_size": files_info[0].file_size if files_info and len(files_info) > 0 else None,
            "file_type": files_info[0].file_type if files_info and len(files_info) > 0 else None,
            "change_description": revision.change_description if revision else None,
            "created_at": revision.created_at if revision else None,
            "uploaded_by": revision.uploaded_by if revision else None,
            "current_step": step_info,
            "current_description": description_info,
            "sequence_order": sequence.sequence_order if sequence else None,
            "is_final": sequence.is_final if sequence else None,
            "requires_transmittal": sequence.requires_transmittal if sequence else None,
            # Дата/время выпуска (создания ревизии)
            "release_date": revision.created_at.isoformat() if revision and revision.created_at else None,
            # Due date вычисляется как created_at + due_days
            "due_date": due_date.isoformat() if due_date else None,
            # Количество дней на выполнение
            "due_days": sequence.due_days if sequence and sequence.due_days else None,
            # Флаг просроченности
            "is_overdue": is_overdue,
            # Компания, от которой ожидается ответ (из входящего трансмиттала)
            "awaiting_company": awaiting_company_by_revision_id.get(revision.id) if revision else None
        })
    
    return result


@router.post("/approve/{document_id}")
async def approve_document(
    document_id: int,
    request: ApproveRequest,
    http_request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    comments = request.comments
    """
    Утвердить документ (внутреннее утверждение)
    """
    # Получаем документ
    document = db.query(Document).filter(Document.id == document_id).first()
    if not document:
        raise HTTPException(status_code=404, detail="Документ не найден")
    
    # Получаем последнюю ревизию
    latest_revision = db.query(DocumentRevision).filter(
        DocumentRevision.document_id == document_id,
        DocumentRevision.is_deleted == 0
    ).order_by(DocumentRevision.created_at.desc()).first()
    
    if not latest_revision:
        raise HTTPException(status_code=404, detail="Ревизия документа не найдена")
    
    # Проверяем, требует ли документ трансмиттал
    from app.models.project import WorkflowPresetSequence, Project
    workflow_sequence = db.query(WorkflowPresetSequence).join(
        Project, Project.workflow_preset_id == WorkflowPresetSequence.preset_id
    ).join(
        Document, Document.project_id == Project.id
    ).filter(
        Document.id == document_id,
        WorkflowPresetSequence.revision_description_id == latest_revision.revision_description_id,
        WorkflowPresetSequence.revision_step_id == latest_revision.revision_step_id
    ).first()
    
    if workflow_sequence and workflow_sequence.requires_transmittal:
        raise HTTPException(
            status_code=400, 
            detail="Документ должен быть утвержден через трансмиттал"
        )
    
    # Проверяем, не находится ли ревизия в активном трансмиттале
    from app.models.transmittal import TransmittalRevision, Transmittal
    from app.models.references import TransmittalStatus
    active_transmittal = db.query(TransmittalRevision).join(
        Transmittal, Transmittal.id == TransmittalRevision.transmittal_id
    ).join(
        TransmittalStatus, TransmittalStatus.id == Transmittal.status_id
    ).filter(
        TransmittalRevision.revision_id == latest_revision.id,
        TransmittalStatus.name.in_(["Draft", "Sent"])  # Активные статусы трансмиттала
    ).first()
    
    if active_transmittal:
        raise HTTPException(
            status_code=400,
            detail=f"Ревизия находится в активном трансмиттале #{active_transmittal.transmittal.transmittal_number}"
        )
    
    # Получаем статус "Approved"
    approved_status = db.query(WorkflowStatus).filter(
        WorkflowStatus.name == "Approved"
    ).first()
    
    if not approved_status:
        raise HTTPException(status_code=500, detail="Статус 'Approved' не найден")
    
    # Сохраняем старый статус перед обновлением
    old_status_id = latest_revision.workflow_status_id
    
    # Валидируем переход статусов
    from app.utils.workflow_status_validator import WorkflowStatusValidator
    if not WorkflowStatusValidator.validate_transition(db, old_status_id, approved_status.id):
        from_status_name = latest_revision.workflow_status.name if latest_revision.workflow_status else "Draft"
        error_msg = WorkflowStatusValidator.get_transition_error_message(from_status_name, "Approved")
        raise HTTPException(status_code=400, detail=error_msg)
    
    # Обновляем статус ревизии
    latest_revision.workflow_status_id = approved_status.id
    
    # Создаем запись в document_workflow_history
    from app.models.document_workflow_history import DocumentWorkflowHistory
    # Используем комментарий только если он не пустой
    final_comments = comments if comments and comments.strip() else None
    workflow_history = DocumentWorkflowHistory(
        revision_id=latest_revision.id,
        from_status_id=old_status_id,
        to_status_id=approved_status.id,
        user_id=current_user.id,
        action_type="approve",
        comments=final_comments
    )
    db.add(workflow_history)
    db.commit()
    
    # Проверяем и обновляем статус проекта при первом утвержденном документе
    _check_and_update_project_status_on_approval(document.project_id, db)
    
    # Логирование действия
    old_values = {
        "id": latest_revision.id,
        "document_id": document_id,
        "workflow_status_id": old_status_id,
    }
    new_values = {
        "id": latest_revision.id,
        "document_id": document_id,
        "workflow_status_id": approved_status.id,
        "action": "approve",
    }
    log_action(
        db=db,
        user_id=current_user.id,
        action="approve",
        entity_type="document",
        entity_id=document_id,
        old_values=old_values,
        new_values=new_values,
        request=http_request,
    )
    
    return {
        "message": "Документ утвержден",
        "document_id": document_id,
        "revision_id": latest_revision.id,
        "approved_by": current_user.id
    }


@router.post("/reject/{document_id}")
async def reject_document(
    document_id: int,
    request: RejectRequest,
    http_request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    comments = request.comments
    print(f"DEBUG: reject_document called with document_id={document_id}, comments='{comments}'")
    """
    Отклонить документ (внутреннее отклонение)
    """
    # Получаем документ
    document = db.query(Document).filter(Document.id == document_id).first()
    if not document:
        raise HTTPException(status_code=404, detail="Документ не найден")
    
    # Получаем последнюю ревизию
    latest_revision = db.query(DocumentRevision).filter(
        DocumentRevision.document_id == document_id,
        DocumentRevision.is_deleted == 0
    ).order_by(DocumentRevision.created_at.desc()).first()
    
    if not latest_revision:
        raise HTTPException(status_code=404, detail="Ревизия документа не найдена")
    
    # Проверяем, требует ли документ трансмиттал
    from app.models.project import WorkflowPresetSequence, Project
    workflow_sequence = db.query(WorkflowPresetSequence).join(
        Project, Project.workflow_preset_id == WorkflowPresetSequence.preset_id
    ).join(
        Document, Document.project_id == Project.id
    ).filter(
        Document.id == document_id,
        WorkflowPresetSequence.revision_description_id == latest_revision.revision_description_id,
        WorkflowPresetSequence.revision_step_id == latest_revision.revision_step_id
    ).first()
    
    if workflow_sequence and workflow_sequence.requires_transmittal:
        raise HTTPException(
            status_code=400, 
            detail="Документ должен быть отклонен через трансмиттал"
        )
    
    # Проверяем, не находится ли ревизия в активном трансмиттале
    from app.models.transmittal import TransmittalRevision, Transmittal
    from app.models.references import TransmittalStatus
    active_transmittal = db.query(TransmittalRevision).join(
        Transmittal, Transmittal.id == TransmittalRevision.transmittal_id
    ).join(
        TransmittalStatus, TransmittalStatus.id == Transmittal.status_id
    ).filter(
        TransmittalRevision.revision_id == latest_revision.id,
        TransmittalStatus.name.in_(["Draft", "Sent"])  # Активные статусы трансмиттала
    ).first()
    
    if active_transmittal:
        raise HTTPException(
            status_code=400,
            detail=f"Ревизия находится в активном трансмиттале #{active_transmittal.transmittal.transmittal_number}"
        )
    
    # Получаем статус "Rejected"
    rejected_status = db.query(WorkflowStatus).filter(
        WorkflowStatus.name == "Rejected"
    ).first()
    
    if not rejected_status:
        raise HTTPException(status_code=500, detail="Статус 'Rejected' не найден")
    
    # Сохраняем старый статус перед обновлением
    old_status_id = latest_revision.workflow_status_id
    
    # Валидируем переход статусов
    from app.utils.workflow_status_validator import WorkflowStatusValidator
    if not WorkflowStatusValidator.validate_transition(db, old_status_id, rejected_status.id):
        from_status_name = latest_revision.workflow_status.name if latest_revision.workflow_status else "Draft"
        error_msg = WorkflowStatusValidator.get_transition_error_message(from_status_name, "Rejected")
        raise HTTPException(status_code=400, detail=error_msg)
    
    # Обновляем статус ревизии
    latest_revision.workflow_status_id = rejected_status.id
    
    # Создаем запись в document_workflow_history
    from app.models.document_workflow_history import DocumentWorkflowHistory
    # Используем комментарий только если он не пустой
    final_comments = comments if comments and comments.strip() else None
    print(f"DEBUG: Creating reject workflow_history with comments='{final_comments}'")
    workflow_history = DocumentWorkflowHistory(
        revision_id=latest_revision.id,
        from_status_id=old_status_id,
        to_status_id=rejected_status.id,
        user_id=current_user.id,
        action_type="reject",
        comments=final_comments
    )
    db.add(workflow_history)
    db.commit()
    print(f"DEBUG: reject workflow_history created with id={workflow_history.id}")
    
    # Логирование действия
    old_values = {
        "id": latest_revision.id,
        "document_id": document_id,
        "workflow_status_id": old_status_id,
    }
    new_values = {
        "id": latest_revision.id,
        "document_id": document_id,
        "workflow_status_id": rejected_status.id,
        "action": "reject",
    }
    log_action(
        db=db,
        user_id=current_user.id,
        action="reject",
        entity_type="document",
        entity_id=document_id,
        old_values=old_values,
        new_values=new_values,
        request=http_request,
    )
    
    return {
        "message": "Документ отклонен",
        "document_id": document_id,
        "revision_id": latest_revision.id,
        "rejected_by": current_user.id,
        "comments": comments
    }


@router.get("/stats")
async def get_reviews_stats(
    project_id: int = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Получить статистику по ревью для дашборда:
    - Всего ожидающих ответа
    - Внутренние (requires_transmittal = False)
    - Через трансмиттал (requires_transmittal = True)
    - Просроченные (due_date < текущая дата)
    """
    # Получаем статус "In Review"
    in_review_status = db.query(WorkflowStatus).filter(
        WorkflowStatus.name == "In Review"
    ).first()
    
    if not in_review_status:
        return {
            "total": 0,
            "internal": 0,
            "transmittal": 0,
            "overdue": 0
        }
    
    # Создаем подзапрос для получения последней ревизии каждого документа (без фильтра по статусу)
    # Важно: сначала находим последнюю ревизию, потом проверяем её статус
    latest_revision_subquery = db.query(
        DocumentRevision.document_id,
        func.max(DocumentRevision.created_at).label('max_created_at')
    ).filter(
        DocumentRevision.is_deleted == 0
    ).group_by(DocumentRevision.document_id).subquery()
    
    # Базовый запрос для всех ревью в статусе "In Review"
    # Ищем документы, у которых последняя ревизия в статусе "In Review"
    base_query = db.query(
        Document,
        DocumentRevision,
        WorkflowPresetSequence
    ).join(
        latest_revision_subquery,
        Document.id == latest_revision_subquery.c.document_id
    ).join(
        DocumentRevision,
        and_(
            DocumentRevision.document_id == Document.id,
            DocumentRevision.created_at == latest_revision_subquery.c.max_created_at,
            DocumentRevision.is_deleted == 0,
            DocumentRevision.workflow_status_id == in_review_status.id  # Проверяем статус последней ревизии
        )
    ).join(
        Project,
        Project.id == Document.project_id
    ).outerjoin(
        WorkflowPresetSequence,
        and_(
            WorkflowPresetSequence.preset_id == Project.workflow_preset_id,
            WorkflowPresetSequence.revision_step_id == DocumentRevision.revision_step_id,
            WorkflowPresetSequence.revision_description_id == DocumentRevision.revision_description_id
        )
    ).filter(
        Document.is_deleted == 0
    )
    
    if project_id:
        base_query = base_query.filter(Project.id == project_id)
    
    # Получаем все результаты
    results = base_query.all()
    
    # Текущая дата для проверки просроченных
    now = datetime.now(timezone.utc)
    
    total = 0
    internal = 0
    transmittal = 0
    overdue = 0
    
    for row in results:
        doc, revision, sequence = row
        # revision не может быть None, так как мы используем join, но на всякий случай проверяем
        if not revision:
            continue
        
        total += 1
        
        # Проверяем requires_transmittal
        requires_transmittal = sequence.requires_transmittal if sequence else False
        
        if requires_transmittal:
            transmittal += 1
        else:
            internal += 1
        
        # Проверяем просроченные (due_date = created_at + due_days)
        if sequence and sequence.due_days and revision.created_at:
            due_date = revision.created_at + timedelta(days=sequence.due_days)
            if due_date < now:
                overdue += 1
    
    return {
        "total": total,
        "internal": internal,
        "transmittal": transmittal,
        "overdue": overdue
    }


@router.get("/export-excel")
async def export_reviews_to_excel(
    project_id: Optional[int] = Query(None),
    search: Optional[str] = Query(None),
    selected_company: Optional[str] = Query(None),
    only_overdue: Optional[bool] = Query(False),
    language: Optional[str] = Query("ru"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Экспорт ревью в Excel файл
    """
    # Получаем данные ревью (используем ту же логику, что и в get_pending_approvals)
    in_review_status = db.query(WorkflowStatus).filter(
        WorkflowStatus.name == "In Review"
    ).first()
    
    if not in_review_status:
        raise HTTPException(status_code=404, detail="Статус 'In Review' не найден")
    
    # Создаем подзапрос для получения последней ревизии каждого документа
    latest_revision_subquery = db.query(
        DocumentRevision.document_id,
        func.max(DocumentRevision.created_at).label('max_created_at')
    ).group_by(DocumentRevision.document_id).subquery()
    
    # Запрос с информацией о последовательности
    query = db.query(
        Document,
        DocumentRevision,
        Project,
        WorkflowPresetSequence
    ).outerjoin(
        latest_revision_subquery,
        Document.id == latest_revision_subquery.c.document_id
    ).outerjoin(
        DocumentRevision,
        and_(
            DocumentRevision.document_id == Document.id,
            DocumentRevision.created_at == latest_revision_subquery.c.max_created_at
        )
    ).join(
        Project,
        Project.id == Document.project_id
    ).outerjoin(
        WorkflowPresetSequence,
        and_(
            WorkflowPresetSequence.preset_id == Project.workflow_preset_id,
            WorkflowPresetSequence.revision_step_id == DocumentRevision.revision_step_id,
            WorkflowPresetSequence.revision_description_id == DocumentRevision.revision_description_id
        )
    ).filter(
        Document.is_deleted == 0,
        DocumentRevision.workflow_status_id == in_review_status.id
    )
    
    if project_id:
        query = query.filter(Project.id == project_id)
    
    # Получаем все результаты (без пагинации для экспорта)
    results = query.order_by(Document.updated_at.desc()).all()
    
    # Загружаем все нужные данные (аналогично get_pending_approvals)
    revision_step_ids = []
    revision_description_ids = []
    
    for row in results:
        doc, revision, project, sequence = row
        if revision and revision.revision_step_id:
            revision_step_ids.append(revision.revision_step_id)
        if revision and revision.revision_description_id:
            revision_description_ids.append(revision.revision_description_id)
    
    revision_steps = {}
    if revision_step_ids:
        for rs in db.query(RevisionStep).filter(RevisionStep.id.in_(revision_step_ids)).all():
            revision_steps[rs.id] = rs
    
    revision_descriptions = {}
    if revision_description_ids:
        for rd in db.query(RevisionDescription).filter(RevisionDescription.id.in_(revision_description_ids)).all():
            revision_descriptions[rd.id] = rd
    
    from collections import defaultdict
    revision_ids = [row[1].id for row in results if row[1] is not None]
    files_by_revision_id = defaultdict(list)
    if revision_ids:
        all_files = db.query(FileModel).filter(
            FileModel.revision_id.in_(revision_ids),
            FileModel.is_deleted == 0
        ).all()
        for f in all_files:
            files_by_revision_id[f.revision_id].append(f)
    
    # Загружаем информацию о компаниях-получателях
    from app.models.transmittal import Transmittal, TransmittalRevision
    from app.models.references import Company, TransmittalStatus
    
    awaiting_company_by_revision_id = {}
    if revision_ids:
        sent_status = db.query(TransmittalStatus).filter(
            TransmittalStatus.name == "Sent"
        ).first()
        
        if sent_status:
            transmittal_data = db.query(
                TransmittalRevision.revision_id,
                Company.id,
                Company.name,
                Company.name_native
            ).join(
                Transmittal, Transmittal.id == TransmittalRevision.transmittal_id
            ).join(
                Company, Company.id == Transmittal.counterparty_id
            ).filter(
                TransmittalRevision.revision_id.in_(revision_ids),
                Transmittal.direction == 'out',
                Transmittal.status_id == sent_status.id,
                Transmittal.is_deleted == 0
            ).all()
            
            for rev_id, company_id, company_name, company_name_native in transmittal_data:
                awaiting_company_by_revision_id[rev_id] = {
                    "id": company_id,
                    "name": company_name,
                    "name_native": company_name_native
                }
    
    # Формируем данные для Excel
    excel_rows = []
    now = datetime.now(timezone.utc)
    
    # Вспомогательные функции для форматирования
    def format_date(date_obj):
        if not date_obj:
            return ''
        if isinstance(date_obj, str):
            try:
                date_obj = datetime.fromisoformat(date_obj.replace('Z', '+00:00'))
            except:
                return ''
        return date_obj.strftime('%d.%m.%Y %H:%M')
    
    def format_file_size(bytes_size):
        if not bytes_size:
            return '0 B'
        k = 1024
        sizes = ['B', 'KB', 'MB', 'GB']
        i = 0
        size = float(bytes_size)
        while size >= k and i < len(sizes) - 1:
            size /= k
            i += 1
        return f"{round(size * 100) / 100} {sizes[i]}"
    
    # Локализация заголовков
    headers_ru = {
        "document": "Документ",
        "title": "Название",
        "project": "Проект",
        "revision": "Ревизия",
        "current_step": "Текущий шаг",
        "awaiting_company": "Ожидается от",
        "release_date": "Дата выпуска",
        "due_days": "Срок (дней)",
        "due_date": "Срок выполнения",
        "overdue": "Просрочено",
        "file": "Файл",
        "size": "Размер"
    }
    
    headers_en = {
        "document": "Document",
        "title": "Title",
        "project": "Project",
        "revision": "Revision",
        "current_step": "Current Step",
        "awaiting_company": "Awaiting Company",
        "release_date": "Release Date",
        "due_days": "Due Days",
        "due_date": "Due Date",
        "overdue": "Overdue",
        "file": "File",
        "size": "Size"
    }
    
    headers = headers_ru if language == "ru" else headers_en
    
    for row in results:
        doc, revision, project, sequence = row
        
        # Применяем фильтры
        if search:
            search_lower = search.lower()
            if not (search_lower in (doc.title or "").lower() or 
                   search_lower in (doc.number or "").lower() or
                   search_lower in (project.name or "").lower()):
                continue
        
        if only_overdue:
            if revision and revision.created_at and sequence and sequence.due_days:
                due_date = revision.created_at + timedelta(days=sequence.due_days)
                if due_date >= now:
                    continue
            else:
                continue
        
        if selected_company:
            if selected_company == '__internal__':
                if sequence and sequence.requires_transmittal != False:
                    continue
            else:
                awaiting_company = awaiting_company_by_revision_id.get(revision.id if revision else None)
                if not awaiting_company or awaiting_company.get("name") != selected_company:
                    continue
        
        # Получаем информацию о шаге и описании
        step_info = None
        description_info = None
        
        if revision:
            step = revision_steps.get(revision.revision_step_id)
            description = revision_descriptions.get(revision.revision_description_id)
            
            step_info = {
                "id": step.id if step else None,
                "code": step.code if step else None,
                "description": step.description if step else None,
                "description_native": step.description_native if step else None
            } if step else None
            
            description_info = {
                "id": description.id if description else None,
                "code": description.code if description else None,
                "description": description.description if description else None,
                "description_native": description.description_native if description else None
            } if description else None
        
        files_info = files_by_revision_id.get(revision.id, []) if revision else []
        
        # Вычисляем due_date и is_overdue
        due_date = None
        is_overdue = False
        
        if revision and revision.created_at and sequence and sequence.due_days:
            due_date = revision.created_at + timedelta(days=sequence.due_days)
            is_overdue = due_date < now
        
        # Формируем поле просрочено
        overdue_value = ''
        if is_overdue and due_date:
            overdue_days = (now - due_date).days
            if overdue_days > 0:
                if language == "ru":
                    # Склонение для русского
                    last_digit = overdue_days % 10
                    last_two = overdue_days % 100
                    if 11 <= last_two <= 14:
                        days_word = "дней"
                    elif last_digit == 1:
                        days_word = "день"
                    elif 2 <= last_digit <= 4:
                        days_word = "дня"
                    else:
                        days_word = "дней"
                    overdue_value = f"Просрочено на {overdue_days} {days_word}"
                else:
                    days_word = "day" if overdue_days == 1 else "days"
                    overdue_value = f"Overdue by {overdue_days} {days_word}"
            else:
                overdue_value = "Просрочено" if language == "ru" else "Overdue"
        
        # Формируем поле текущего шага
        current_step_value = ''
        if step_info:
            step_desc = (step_info.get("description_native") if language == "ru" else step_info.get("description")) or step_info.get("description") or ''
            if step_desc and step_desc.strip():
                current_step_value = f"{step_info.get('code')} - {step_desc}"
            else:
                current_step_value = step_info.get('code') or ''
        
        # Формируем поле ревизии
        revision_value = ''
        if description_info:
            revision_value = f"{description_info.get('code')}{revision.number if revision else ''}"
        elif revision:
            revision_value = revision.number or ''
        
        # Формируем поле ожидается от
        awaiting_company_value = ''
        awaiting_company = awaiting_company_by_revision_id.get(revision.id if revision else None)
        if awaiting_company:
            awaiting_company_value = awaiting_company.get("name") or ''
        elif sequence and sequence.requires_transmittal == False:
            awaiting_company_value = "Внутреннее ревью" if language == "ru" else "Internal Review"
        
        excel_rows.append([
            doc.number or '',
            doc.title or '',
            project.name if project else '',
            revision_value,
            current_step_value,
            awaiting_company_value,
            format_date(revision.created_at if revision else None),
            sequence.due_days if sequence else '',
            format_date(due_date) if due_date else '',
            overdue_value,
            files_info[0].file_name if files_info and len(files_info) > 0 else '',
            format_file_size(files_info[0].file_size if files_info and len(files_info) > 0 else None)
        ])
    
    # Создаем Excel файл
    wb = Workbook()
    ws = wb.active
    ws.title = "Reviews" if language == "en" else "Ревью"
    
    # Добавляем заголовки
    header_row = [
        headers["document"],
        headers["title"],
        headers["project"],
        headers["revision"],
        headers["current_step"],
        headers["awaiting_company"],
        headers["release_date"],
        headers["due_days"],
        headers["due_date"],
        headers["overdue"],
        headers["file"],
        headers["size"]
    ]
    ws.append(header_row)
    
    # Применяем стили к заголовкам
    # Используем цвет хедера приложения #1976d2
    header_fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Добавляем данные
    data_alignment = Alignment(vertical="center", wrap_text=True)
    data_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0')
    )
    
    # Создаем стиль для просроченных записей (красный текст)
    overdue_font = Font(color="FF0000", bold=False)  # Красный цвет
    
    # Определяем индекс колонки "Overdue" из заголовков
    overdue_col_index = None
    for idx, header in enumerate(header_row, start=1):
        if language == "ru":
            if header == headers["overdue"]:  # "Просрочено"
                overdue_col_index = idx
                break
        else:
            if header == headers["overdue"]:  # "Overdue"
                overdue_col_index = idx
                break
    
    for row_data in excel_rows:
        ws.append(row_data)
        for col_idx, cell in enumerate(ws[ws.max_row], start=1):
            cell.alignment = data_alignment
            cell.border = data_border
            
            # Применяем красный цвет к колонке "Overdue", если значение не пустое
            if overdue_col_index and col_idx == overdue_col_index:
                cell_value = str(cell.value) if cell.value else ''
                if cell_value and (cell_value != ''):
                    cell.font = overdue_font
    
    # Настраиваем ширину колонок
    column_widths = [15, 30, 20, 12, 25, 20, 20, 12, 20, 20, 25, 12]
    for idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    # Устанавливаем высоту строки для заголовков
    ws.row_dimensions[1].height = 25
    
    # Замораживаем первую строку
    ws.freeze_panes = 'A2'
    
    # Добавляем автофильтр для заголовков
    if excel_rows:
        # Определяем диапазон для автофильтра (заголовки + данные)
        last_col = get_column_letter(len(header_row))
        ws.auto_filter.ref = f"A1:{last_col}{len(excel_rows) + 1}"
    
    # Сохраняем в BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Генерируем имя файла
    project_name = "all_projects"
    if project_id:
        project_obj = db.query(Project).filter(Project.id == project_id).first()
        if project_obj:
            project_name = project_obj.name or "all_projects"
    filename = f"reviews_{project_name}_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    
    # Правильно кодируем имя файла для заголовка Content-Disposition
    # Используем percent-encoding для UTF-8 (RFC 5987)
    # Сначала кодируем в UTF-8 байты, затем применяем percent-encoding
    import urllib.parse
    # Правильный способ: кодируем строку в UTF-8, затем применяем quote
    encoded_filename = urllib.parse.quote(filename, safe='', encoding='utf-8')
    
    # Возвращаем файл
    return Response(
        content=output.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"
        }
    )